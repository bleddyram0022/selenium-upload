import openpyxl

path="Documents\Book2.Xlsl"

workbook=openpyxl.load_workbook(path)

sheet=workbook.active

sheet.max_row
sheet.max_column

print(rows)
print(column


def test_yopmail_id_activation(setup):

    driver.get("https://yopmail.com/en/wm")
    driver.implicitly_wait(10)
    driver.maximize_window()
    driver.execute_script("window.scrollBy(0,300)","")
    time.sleep(2)
    driver.find_element(By.XPATH,'//*[@id="mail"]/div/table/tbody/tr[2]/td/table/tbody/tr[1]/td/p[4]/a').click()
    driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/main/div/div/div[2]/div/div/div/div/form/div[1]/div/input').send_keys("dopie1.com")
    driver.find_element(By.XPATH,'//*[@id="dz-password"]').send_keys("Tester@123")
    driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/main/div/div/div[2]/div/div/div/div/form/div[4]').click()


@pytest.fixture()
def test_Admin():
    global driver
    driver = webdriver.Chrome()
    driver.get("https://dopamineexchcontrols.wealwindemo.com/login") #Admin webpage
    driver.maximize_window()
    driver.find_element(By.XPATH,'//*[@id="root"]/div[2]/div/div[2]/div/div/div/div/form/div[1]/div/input').send_keys("dopamine@admin.com")
    driver.find_element(By.XPATH,'//*[@id="root"]/div[2]/div/div[2]/div/div/div/div/form/div[2]/div/input').send_keys("Test@123")
    driver.find_element(By.CLASS_NAME,"icon").click()
    driver.find_element(By.XPATH,'//*[@id="root"]/div[2]/div/div[2]/div/div/div/div/form/button').click()



@pytest.fixture()
def mynewuser():
    global driver
    driver = webdriver.Chrome()
    driver.get("https://dopamineexch.wealwindemo.com/")
    driver.maximize_window()
    driver.find_element(By.XPATH,'//*[@id="navbarTogglerDemo03"]/ul/li[4]/div/div/button[1]').click()
    driver.get("https://dopamineexch.wealwindemo.com/signin")
    driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/main/div/div/div[2]/div/div/div/div/form/div[1]/div/input').send_keys("mynew@yopmail.com")
    driver.find_element(By.XPATH,'//*[@id="dz-password"]').send_keys("Test@123")
    driver.find_element(By.XPATH,'//*[@id="__next"]/div[2]/main/div/div/div[2]/div/div/div/div/form/div[4]').click()


def test_userdeposit(dopie):

  driver.get("https://dopamineexchcontrols.wealwindemo.com/dashboard")

  time.sleep(2)

  driver.find_element(By.XPATH, '//*[@id="root"]/div[2]/div[2]/ul/div/div[1]/div[2]/div/div/div/li[2]/a').click()

  time.sleep(2)

  driver.find_element(By.XPATH,'//*[@id="root"]/div[2]/div[2]/ul/div/div[1]/div[2]/div/div/div/li[2]/ul/li[1]/a').click()

  time.sleep(2)

  driver.get("https://dopamineexchcontrols.wealwindemo.com/user-list")

  time.sleep(2)

  svg_element = driver.find_element(By.CLASS_NAME, "icon")

  time.sleep(3)

  #driver.find_element(By.CLASS_NAME,"icon").click()

  #time.sleep(3)

  driver.get("https://dopamineexchcontrols.wealwindemo.com/userAsset/U2FsdGVkX1xMl3JkdwPON58zYPjyp1SZTp5ec6WSZGEJUPKa20JoeO2pWAl9avV3GsGyV#")

  time.sleep(3)

  #BNB Deposit button click
  button = WebDriverWait(driver, 10).until(
      EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[2]/div[3]/div[2]/div/div/div[2]/div/table/tbody/tr[1]/td[5]/div/button[1]'))
  )
  button.click()

  time.sleep(3)

  driver.find_element(By.XPATH,'/html/body/div[3]/div/div/div/div[2]/input').send_keys("600")

  time.sleep(2)



  button_element = driver.find_element(By.CLASS_NAME, "attach-btn")

  # Click on the update  button element
  button_element.click()

  time.sleep(3)

  driver.get("https://dopamineexchcontrols.wealwindemo.com/userAsset/U2FsdGVkX1xMl3JkdwPON58zYPjyp1SZTp5ec6WSZGEJUPKa20JoeO2pWAl9avV3GsGyV#")

  time.sleep(5)

  #click USDT deposit button

  button = WebDriverWait(driver, 10).until(
      EC.element_to_be_clickable(
          (By.XPATH, '//*[@id="root"]/div[2]/div[3]/div[2]/div/div/div[2]/div/table/tbody/tr[1]/td[5]/div/button[1]'))
  )
  button.click()

  time.sleep(3)

  driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div[2]/input').send_keys("600")

  time.sleep(2)

  button_element = driver.find_element(By.CLASS_NAME, "attach-btn")

  # Click on the button element
  button_element.click()

  time.sleep(3)

  driver.get(
      "https://dopamineexchcontrols.wealwindemo.com/userAsset/U2FsdGVkX1xMl3JkdwPON58zYPjyp1SZTp5ec6WSZGEJUPKa20JoeO2pWAl9avV3GsGyV#")

  time.sleep(5)

  #my new Deposit

  svg_element = driver.find_element(By.CLASS_NAME, "icon")

  time.sleep(3)

  # driver.find_element(By.CLASS_NAME,"icon").click()

  # time.sleep(3)

  driver.get(
      "https://dopamineexchcontrols.wealwindemo.com/userAsset/U2FsdGVkX1xMl3JkdwPON58zYPjyp1SZTp5ec6WSZGEJUPKa20JoeO2pWAl9avV3GsGyV#")

  time.sleep(3)

  #BNB Deposit

  button = WebDriverWait(driver, 10).until(
      EC.element_to_be_clickable(
          (By.XPATH, '//*[@id="root"]/div[2]/div[3]/div[2]/div/div/div[2]/div/table/tbody/tr[1]/td[5]/div/button[1]'))
  )
  button.click()

  time.sleep(3)

  driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div[2]/input').send_keys("600")

  time.sleep(2)

  button_element = driver.find_element(By.CLASS_NAME, "attach-btn")

  # Click on the button element
  button_element.click()

  time.sleep(3)

  driver.get(
      "https://dopamineexchcontrols.wealwindemo.com/userAsset/U2FsdGVkX1xMl3JkdwPON58zYPjyp1SZTp5ec6WSZGEJUPKa20JoeO2pWAl9avV3GsGyV#")

  time.sleep(5)

  #USDT Deposit

  button = WebDriverWait(driver, 10).until(
      EC.element_to_be_clickable(
          (By.XPATH, '//*[@id="root"]/div[2]/div[3]/div[2]/div/div/div[2]/div/table/tbody/tr[1]/td[5]/div/button[1]'))
  )
  button.click()

  time.sleep(3)

  driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div[2]/input').send_keys("600")

  time.sleep(2)

  button_element = driver.find_element(By.CLASS_NAME, "attach-btn")

  # Click on the button element
  button_element.click()

  time.sleep(3)

  driver.get(
      "https://dopamineexchcontrols.wealwindemo.com/userAsset/U2FsdGVkX1xMl3JkdwPON58zYPjyp1SZTp5ec6WSZGEJUPKa20JoeO2pWAl9avV3GsGyV#")

time.sleep(5)


def test_NoEntertheValue(setup):
    driver.get("https://dopamineexch.wealwindemo.com/signup")

    driver.find_element(By.XPATH, '//*[@id="email"]').send_keys("  ")
    driver.find_element(By.XPATH, '//*[@id="password"]').send_keys("  ")

    driver.execute_script("window.scrollBy(0,300)", "")

    driver.find_element(By.XPATH, '//*[@id="confirmPassword"]').send_keys("  ")
    driver.find_element(By.XPATH,
                        '//*[@id="__next"]/div[2]/main/div/div/div[2]/div/div/div/div/form/div[3]/div/span').click()
    driver.find_element(By.XPATH,
                        '//*[@id="__next"]/div[2]/main/div/div/div[2]/div/div/div/div/form/div[5]/button').click()
    time.sleep(5)

